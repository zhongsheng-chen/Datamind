# datamind/core/ml/export/scorecard_exporter.py
"""评分卡导出器

将评分卡（分箱 + WOE + 系数 + 分数）导出为 Excel 文件。

核心功能：
  - export: 导出完整评分卡
  - export_with_summary: 导出评分卡 + 模型摘要 + 分箱统计

特性：
  - 银行标准格式：特征、分箱、WOE、系数、分数
  - 多 Sheet：Scorecard、Summary、Bin Statistics
  - 支持 Reason Code 映射
  - 自动调整列宽

使用示例：
  >>> from datamind.core.ml.export.scorecard_exporter import ScorecardExporter
  >>>
  >>> exporter = ScorecardExporter(reason_engine=reason_engine)
  >>> exporter.export(model, binning, "scorecard.xlsx", factor=20.0, metadata=metadata)
"""

import logging
from typing import Dict, List, Any, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from datamind.core.ml.features.binning import Bin


class ScorecardExporter:
    """评分卡导出器"""

    def __init__(self, reason_engine: Optional[Any] = None):
        """
        初始化导出器

        参数:
            reason_engine: ReasonCodeEngine 实例（可选，用于导出拒绝原因）
        """
        self.reason_engine = reason_engine
        self._logger = logging.getLogger(__name__)

    def export(
        self,
        model,
        binning: Dict[str, List[Bin]],
        filepath: str,
        factor: float,
        metadata: Optional[Dict] = None
    ) -> str:
        """
        导出评分卡到 Excel

        参数:
            model: 模型对象（需支持 get_coef 方法）
            binning: 分箱配置 {feature: [Bin, Bin, ...]}
            filepath: 导出文件路径
            factor: 评分因子
            metadata: 模型元数据（版本、AUC等）

        返回:
            导出文件路径
        """
        wb = Workbook()

        # Sheet1: Scorecard
        self._create_scorecard_sheet(wb, model, binning, factor)

        # Sheet2: Model Summary
        if metadata:
            self._create_summary_sheet(wb, metadata)

        # Sheet3: Bin Statistics
        self._create_bin_stats_sheet(wb, binning)

        # 保存
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

        self._logger.info(f"Scorecard exported to {filepath}")
        return str(path)

    def _create_scorecard_sheet(self, wb, model, binning, factor):
        """创建评分卡主 Sheet"""
        ws = wb.active
        ws.title = "Scorecard"

        # 表头
        headers = ["Feature", "Bin", "Lower", "Upper", "WOE", "Coefficient", "Score"]
        if self.reason_engine:
            headers.extend(["Reason Code", "Reason Message"])

        ws.append(headers)

        # 样式
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = header_font

        # 数据行
        for feature, bins in binning.items():
            coef = self._get_coefficient(model, feature)

            for b in bins:
                score = -factor * coef * b.woe

                row = [
                    feature,
                    b.label,
                    b.lower,
                    b.upper,
                    round(b.woe, 4),
                    round(coef, 4),
                    round(score, 2)
                ]

                if self.reason_engine:
                    reason = self._get_reason_for_bin(feature, b)
                    row.extend([reason.get("code", ""), reason.get("message", "")])

                ws.append(row)

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _create_summary_sheet(self, wb, metadata: Dict):
        """创建模型摘要 Sheet"""
        ws = wb.create_sheet("Model Summary")

        summary_items = [
            ("Model ID", metadata.get("model_id")),
            ("Model Version", metadata.get("model_version")),
            ("Model Type", metadata.get("model_type")),
            ("Task Type", metadata.get("task_type")),
            ("Framework", metadata.get("framework")),
            ("Base Score", metadata.get("base_score")),
            ("PDO", metadata.get("pdo")),
            ("Odds", metadata.get("odds")),
            ("AUC", metadata.get("auc")),
            ("KS", metadata.get("ks")),
            ("Created At", metadata.get("created_at")),
            ("Created By", metadata.get("created_by")),
        ]

        for key, value in summary_items:
            if value is not None:
                ws.append([key, value])

    def _create_bin_stats_sheet(self, wb, binning):
        """创建分箱统计 Sheet"""
        ws = wb.create_sheet("Bin Statistics")

        ws.append(["Feature", "Bin", "WOE", "Bad Rate", "Sample Ratio"])

        for feature, bins in binning.items():
            for b in bins:
                ws.append([
                    feature,
                    b.label,
                    round(b.woe, 4),
                    round(b.bad_rate, 4) if b.bad_rate else None,
                    round(b.sample_ratio, 4) if b.sample_ratio else None
                ])

    def _get_coefficient(self, model, feature: str) -> float:
        """获取特征系数"""
        if hasattr(model, "get_coef"):
            return model.get_coef(feature)

        if hasattr(model, "coef_map"):
            return model.coef_map.get(feature, 0.0)

        self._logger.warning(f"Cannot get coefficient for {feature}")
        return 0.0

    def _get_reason_for_bin(self, feature: str, bin_obj: Bin) -> Dict:
        """获取分箱对应的拒绝原因"""
        if not self.reason_engine:
            return {}

        # 构造一个假的解释结果用于匹配
        fake_explain = {
            "feature": feature,
            "score": -1.0,
            "bin": {
                "label": bin_obj.label,
                "lower": bin_obj.lower,
                "upper": bin_obj.upper,
                "id": bin_obj.id,
                "is_missing": bin_obj.is_missing
            }
        }

        matched = self.reason_engine._match_reason(fake_explain)
        return matched or {}